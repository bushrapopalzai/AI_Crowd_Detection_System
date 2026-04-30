"""
Advanced Object Detection System with Excel-Only Database & Real-Time Streaming
Features: Live Excel writing, persistent append-only storage, modern async architecture
"""

import cv2
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import threading
import time
from collections import deque, defaultdict
from datetime import datetime, timedelta
import json
import os
import queue
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import asyncio
from concurrent.futures import ThreadPoolExecutor
import logging
from pathlib import Path

# Try to import customtkinter for modern UI
try:
    import customtkinter as ctk
    USE_CTK = True
except ImportError:
    USE_CTK = False
    print("CustomTkinter not found, using standard Tkinter")

if USE_CTK:
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("blue")

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelLiveDatabase:
    """
    Real-time Excel database with append-only architecture
    Handles live data streaming with automatic persistence
    """
    
    def __init__(self, filename="detection_records.xlsx"):
        self.filename = filename
        self.lock = threading.Lock()
        self.executor = ThreadPoolExecutor(max_workers=2)
        self.buffer = []
        self.buffer_size = 10  # Flush every N records
        self.last_flush = time.time()
        self.flush_interval = 5  # Or every 5 seconds
        
        # Ensure file exists with proper structure
        self._initialize_excel()
        
        # Start background flush thread
        self.flush_thread = threading.Thread(target=self._background_flush, daemon=True)
        self.flush_thread.start()
        
    def _initialize_excel(self):
        """Create Excel file with proper structure if not exists"""
        if not os.path.exists(self.filename):
            logger.info(f"Creating new Excel database: {self.filename}")
            wb = Workbook()
            
            # Sheet 1: Live Detections (append-only stream)
            ws_live = wb.active
            ws_live.title = "Live_Detections"
            headers = ["Timestamp", "Date", "Time", "Day", "Session_ID", "Frame_Number", 
                      "Class_Name", "Confidence", "BBox_X1", "BBox_Y1", "BBox_X2", 
                      "BBox_Y2", "FPS", "Source_Type", "Source_Path"]
            ws_live.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            for cell in ws_live[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Freeze header row
            ws_live.freeze_panes = "A2"
            
            # Sheet 2: Session Summary
            ws_sessions = wb.create_sheet("Sessions")
            session_headers = ["Session_ID", "Start_Time", "End_Time", "Date", "Day", 
                             "Source_Type", "Source_Path", "Total_Frames", "Total_Detections",
                             "Unique_Classes", "Avg_Confidence", "Duration_Seconds", "Status"]
            ws_sessions.append(session_headers)
            for cell in ws_sessions[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Sheet 3: Hourly Analytics
            ws_hourly = wb.create_sheet("Hourly_Analytics")
            hourly_headers = ["Timestamp", "Date", "Hour", "Day", "Session_ID", "Class_Name", 
                            "Count", "Avg_Confidence", "Avg_FPS"]
            ws_hourly.append(hourly_headers)
            for cell in ws_hourly[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Sheet 4: Daily Summary (auto-updated)
            ws_daily = wb.create_sheet("Daily_Summary")
            daily_headers = ["Date", "Day", "Total_Sessions", "Total_Detections", 
                           "Unique_Classes", "Most_Detected_Class", "Peak_Hour"]
            ws_daily.append(daily_headers)
            for cell in ws_daily[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Set column widths for readability
            for ws in [ws_live, ws_sessions, ws_hourly, ws_daily]:
                for col in range(1, 20):
                    ws.column_dimensions[get_column_letter(col)].width = 15
            
            wb.save(self.filename)
            logger.info("Excel database initialized successfully")
        else:
            logger.info(f"Connected to existing Excel database: {self.filename}")
    
    def _background_flush(self):
        """Background thread to periodically flush buffer to disk"""
        while True:
            time.sleep(1)
            current_time = time.time()
            if (len(self.buffer) >= self.buffer_size or 
                (current_time - self.last_flush >= self.flush_interval and self.buffer)):
                self._flush_buffer()
    
    def _flush_buffer(self):
        """Write buffered data to Excel file"""
        if not self.buffer:
            return
            
        with self.lock:
            buffer_copy = self.buffer.copy()
            self.buffer = []
            
        try:
            # Load workbook in append mode
            wb = load_workbook(self.filename)
            ws = wb["Live_Detections"]
            
            # Append all buffered records
            for record in buffer_copy:
                row_data = [
                    record['timestamp'],
                    record['date'],
                    record['time'],
                    record['day'],
                    record['session_id'],
                    record['frame_number'],
                    record['class_name'],
                    record['confidence'],
                    record['bbox_x1'],
                    record['bbox_y1'],
                    record['bbox_x2'],
                    record['bbox_y2'],
                    record['fps'],
                    record['source_type'],
                    record['source_path']
                ]
                ws.append(row_data)
            
            # Auto-adjust column widths for new data
            for col in range(1, len(row_data) + 1):
                max_length = 0
                column = get_column_letter(col)
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(self.filename)
            self.last_flush = time.time()
            logger.debug(f"Flushed {len(buffer_copy)} records to Excel")
            
        except Exception as e:
            logger.error(f"Error flushing to Excel: {e}")
            # Put records back in buffer
            with self.lock:
                self.buffer = buffer_copy + self.buffer
    
    def insert_detection(self, session_id, frame_number, detection_data, fps, source_type, source_path):
        """Insert detection immediately (buffered write)"""
        now = datetime.now()
        record = {
            'timestamp': now,
            'date': now.strftime('%Y-%m-%d'),
            'time': now.strftime('%H:%M:%S.%f')[:-3],  # Milliseconds precision
            'day': now.strftime('%A'),
            'session_id': session_id,
            'frame_number': frame_number,
            'class_name': detection_data['class'],
            'confidence': round(detection_data['confidence'], 4),
            'bbox_x1': detection_data['bbox'][0],
            'bbox_y1': detection_data['bbox'][1],
            'bbox_x2': detection_data['bbox'][2],
            'bbox_y2': detection_data['bbox'][3],
            'fps': round(fps, 2),
            'source_type': source_type,
            'source_path': source_path
        }
        
        with self.lock:
            self.buffer.append(record)
        
        # Immediate flush if buffer is getting large
        if len(self.buffer) >= self.buffer_size:
            threading.Thread(target=self._flush_buffer, daemon=True).start()
    
    def start_session(self, session_id, source_type, source_path):
        """Record new session start"""
        now = datetime.now()
        session_data = {
            'session_id': session_id,
            'start_time': now,
            'date': now.strftime('%Y-%m-%d'),
            'day': now.strftime('%A'),
            'source_type': source_type,
            'source_path': source_path,
            'status': 'RUNNING'
        }
        # Store in memory until session ends
        self._active_sessions = getattr(self, '_active_sessions', {})
        self._active_sessions[session_id] = session_data
        logger.info(f"Session {session_id} started")
    
    def end_session(self, session_id, total_frames, total_detections, unique_classes, avg_confidence):
        """Finalize session and write to Sessions sheet"""
        if not hasattr(self, '_active_sessions') or session_id not in self._active_sessions:
            return
            
        session = self._active_sessions[session_id]
        end_time = datetime.now()
        duration = (end_time - session['start_time']).total_seconds()
        
        with self.lock:
            try:
                wb = load_workbook(self.filename)
                ws = wb["Sessions"]
                
                row_data = [
                    session_id,
                    session['start_time'],
                    end_time,
                    session['date'],
                    session['day'],
                    session['source_type'],
                    session['source_path'],
                    total_frames,
                    total_detections,
                    unique_classes,
                    round(avg_confidence, 4) if avg_confidence else 0,
                    round(duration, 2),
                    'COMPLETED'
                ]
                ws.append(row_data)
                
                # Update daily summary
                self._update_daily_summary(wb, session['date'])
                
                wb.save(self.filename)
                del self._active_sessions[session_id]
                logger.info(f"Session {session_id} ended and saved")
                
            except Exception as e:
                logger.error(f"Error ending session: {e}")
    
    def _update_daily_summary(self, wb, date):
        """Update daily statistics"""
        try:
            # Read live data for this date
            ws_live = wb["Live_Detections"]
            daily_data = []
            
            for row in ws_live.iter_rows(min_row=2, values_only=True):
                if row[1] == date:  # Date column
                    daily_data.append(row)
            
            if not daily_data:
                return
            
            # Calculate stats
            total_detections = len(daily_data)
            classes = [row[6] for row in daily_data]  # Class_Name column
            most_common = max(set(classes), key=classes.count) if classes else "N/A"
            
            # Count sessions
            sessions = set(row[4] for row in daily_data)  # Session_ID column
            
            # Find peak hour
            hours = [row[0].hour if isinstance(row[0], datetime) else 0 for row in daily_data]
            peak_hour = max(set(hours), key=hours.count) if hours else 0
            
            # Update or append to Daily_Summary
            ws_daily = wb["Daily_Summary"]
            date_exists = False
            
            for idx, row in enumerate(ws_daily.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] == date:
                    # Update existing
                    ws_daily.cell(row=idx, column=3, value=len(sessions))
                    ws_daily.cell(row=idx, column=4, value=total_detections)
                    ws_daily.cell(row=idx, column=5, value=len(set(classes)))
                    ws_daily.cell(row=idx, column=6, value=most_common)
                    ws_daily.cell(row=idx, column=7, value=f"{peak_hour:02d}:00")
                    date_exists = True
                    break
            
            if not date_exists:
                day_name = datetime.strptime(date, '%Y-%m-%d').strftime('%A')
                ws_daily.append([
                    date, day_name, len(sessions), total_detections,
                    len(set(classes)), most_common, f"{peak_hour:02d}:00"
                ])
                
        except Exception as e:
            logger.error(f"Error updating daily summary: {e}")
    
    def get_live_stats(self):
        """Get current buffer stats and file info"""
        return {
            'buffered_records': len(self.buffer),
            'file_size_mb': round(os.path.getsize(self.filename) / (1024 * 1024), 2) if os.path.exists(self.filename) else 0,
            'last_flush': datetime.fromtimestamp(self.last_flush).strftime('%H:%M:%S')
        }
    
    def force_flush(self):
        """Force immediate write to disk"""
        self._flush_buffer()
    
    def read_all_data(self):
        """Read all detection data as DataFrame"""
        self.force_flush()
        try:
            return pd.read_excel(self.filename, sheet_name="Live_Detections")
        except Exception as e:
            logger.error(f"Error reading Excel: {e}")
            return pd.DataFrame()
    
    def get_date_range_data(self, start_date, end_date):
        """Get data for specific date range"""
        df = self.read_all_data()
        if df.empty:
            return df
        mask = (df['Date'] >= start_date) & (df['Date'] <= end_date)
        return df.loc[mask]
    
    def create_analytics_charts(self):
        """Create charts in Excel for visualization"""
        self.force_flush()
        try:
            wb = load_workbook(self.filename)
            
            # Create charts sheet if not exists
            if "Charts" not in wb.sheetnames:
                ws_charts = wb.create_sheet("Charts")
            else:
                ws_charts = wb["Charts"]
                ws_charts.delete_rows(1, ws_charts.max_row)
            
            # Chart 1: Detections by Class (Pie Chart)
            df = self.read_all_data()
            if not df.empty:
                class_counts = df['Class_Name'].value_counts().head(10)
                
                # Write data for chart
                ws_charts.cell(row=1, column=1, value="Class")
                ws_charts.cell(row=1, column=2, value="Count")
                for idx, (cls, count) in enumerate(class_counts.items(), start=2):
                    ws_charts.cell(row=idx, column=1, value=cls)
                    ws_charts.cell(row=idx, column=2, value=count)
                
                pie = PieChart()
                pie.title = "Top 10 Detected Objects"
                labels = Reference(ws_charts, min_col=1, min_row=2, max_row=len(class_counts)+1)
                data = Reference(ws_charts, min_col=2, min_row=1, max_row=len(class_counts)+1)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                ws_charts.add_chart(pie, "D2")
            
            wb.save(self.filename)
            logger.info("Analytics charts created in Excel")
            
        except Exception as e:
            logger.error(f"Error creating charts: {e}")


class SmartVideoBuffer:
    """
    Intelligent video buffering with frame skipping for performance
    """
    
    def __init__(self, maxsize=5):
        self.queue = queue.Queue(maxsize=maxsize)
        self.frame_count = 0
        self.skip_factor = 1  # Process every Nth frame
        self.target_fps = 30
        
    def put(self, frame, detections, fps):
        """Add frame with intelligent dropping if queue is full"""
        self.frame_count += 1
        
        # Skip frames if queue is getting full
        if self.queue.qsize() >= self.queue.maxsize - 1:
            self.skip_factor = min(self.skip_factor + 1, 5)
        elif self.queue.qsize() < 2:
            self.skip_factor = max(self.skip_factor - 1, 1)
        
        if self.frame_count % self.skip_factor == 0:
            try:
                self.queue.put_nowait((frame, detections, fps, datetime.now()))
                return True
            except queue.Full:
                return False
        return False
    
    def get(self, timeout=0.033):
        """Get frame with timeout"""
        try:
            return self.queue.get(timeout=timeout)
        except queue.Empty:
            return None


class DetectionModel:
    """Optimized YOLO Model wrapper with batch processing support"""
    
    def __init__(self, model_name='yolov8n.pt'):
        self.model_name = model_name
        self.model = None
        self.classes = []
        self.colors = {}
        self.conf_threshold = 0.5
        self.load_model()
        
    def load_model(self):
        try:
            from ultralytics import YOLO
            logger.info(f"Loading {self.model_name}...")
            self.model = YOLO(self.model_name)
            self.classes = self.model.names
            
            # Generate distinct colors for each class
            np.random.seed(42)
            for cls_name in self.classes.values():
                self.colors[cls_name] = tuple(map(int, np.random.randint(0, 255, 3)))
            logger.info(f"Model loaded with {len(self.classes)} classes")
            return True
        except Exception as e:
            logger.error(f"Error loading model: {e}")
            return False
    
    def detect(self, frame, conf_threshold=None):
        """Run detection with optional confidence override"""
        if self.model is None:
            return frame, []
        
        conf = conf_threshold or self.conf_threshold
        
        try:
            # Run inference
            results = self.model(frame, conf=conf, verbose=False)
            detections = []
            
            # Use first result (batch size 1)
            result = results[0]
            
            # Extract detections
            for box in result.boxes:
                x1, y1, x2, y2 = map(int, box.xyxy[0])
                conf_score = float(box.conf[0])
                cls_id = int(box.cls[0])
                cls_name = self.classes[cls_id]
                
                detections.append({
                    'class': cls_name,
                    'confidence': conf_score,
                    'bbox': (x1, y1, x2, y2)
                })
            
            # Draw annotations
            annotated_frame = self._annotate_frame(frame.copy(), detections)
            return annotated_frame, detections
            
        except Exception as e:
            logger.error(f"Detection error: {e}")
            return frame, []
    
    def _annotate_frame(self, frame, detections):
        """Draw bounding boxes and labels"""
        for det in detections:
            x1, y1, x2, y2 = det['bbox']
            cls_name = det['class']
            conf = det['confidence']
            color = self.colors.get(cls_name, (0, 255, 0))
            
            # Draw box
            cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
            
            # Draw label background
            label = f"{cls_name} {conf:.2f}"
            (text_w, text_h), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
            cv2.rectangle(frame, (x1, y1 - text_h - 10), (x1 + text_w, y1), color, -1)
            
            # Draw text
            cv2.putText(frame, label, (x1, y1 - 5), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        
        return frame


class VideoProcessor(threading.Thread):
    """High-performance video processing with database integration"""
    
    def __init__(self, model, buffer, stop_event, db, session_id, source_type, source_path):
        super().__init__(daemon=True)
        self.model = model
        self.buffer = buffer
        self.stop_event = stop_event
        self.db = db
        self.session_id = session_id
        self.source_type = source_type
        self.source_path = source_path
        
        self.cap = None
        self.frame_count = 0
        self.total_detections = 0
        self.all_confidences = []
        self.unique_classes = set()
        
    def run(self):
        self.cap = cv2.VideoCapture(self.source_path)
        if not self.cap.isOpened():
            logger.error("Failed to open video source")
            self.stop_event.set()
            return
        
        # Get video properties
        fps = self.cap.get(cv2.CAP_PROP_FPS) or 30
        logger.info(f"Video FPS: {fps}")
        
        prev_time = time.time()
        frame_time = 1.0 / fps
        
        while not self.stop_event.is_set():
            ret, frame = self.cap.read()
            if not ret:
                logger.info("End of video stream")
                break
            
            current_time = time.time()
            actual_fps = 1.0 / (current_time - prev_time) if (current_time - prev_time) > 0 else 0
            prev_time = current_time
            
            self.frame_count += 1
            
            # Process detection
            annotated_frame, detections = self.model.detect(frame)
            
            # Update statistics
            self.total_detections += len(detections)
            for det in detections:
                self.all_confidences.append(det['confidence'])
                self.unique_classes.add(det['class'])
                
                # Write to Excel immediately (buffered)
                self.db.insert_detection(
                    self.session_id, 
                    self.frame_count, 
                    det, 
                    actual_fps,
                    self.source_type,
                    str(self.source_path)
                )
            
            # Add FPS overlay
            cv2.putText(annotated_frame, f"FPS: {actual_fps:.1f}", (10, 30),
                       cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            cv2.putText(annotated_frame, f"Frame: {self.frame_count}", (10, 70),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            
            # Add to display buffer
            self.buffer.put(annotated_frame, detections, actual_fps)
            
            # Frame rate limiting
            elapsed = time.time() - current_time
            if elapsed < frame_time:
                time.sleep(frame_time - elapsed)
        
        self._cleanup()
    
    def _cleanup(self):
        """Cleanup and final statistics"""
        if self.cap:
            self.cap.release()
        
        avg_conf = sum(self.all_confidences) / len(self.all_confidences) if self.all_confidences else 0
        
        # End session in database
        self.db.end_session(
            self.session_id,
            self.frame_count,
            self.total_detections,
            len(self.unique_classes),
            avg_conf
        )
        
        logger.info(f"Session {self.session_id} completed: {self.frame_count} frames, {self.total_detections} detections")


class ModernDetectionGUI:
    """Modern GUI with real-time Excel integration"""
    
    def __init__(self):
        self.root = ctk.CTk() if USE_CTK else tk.Tk()
        self.root.title("AI Detection System - Excel Live Database")
        self.root.geometry("1600x1000")
        
        # Initialize components
        self.model = DetectionModel('yolov8n.pt')
        self.db = ExcelLiveDatabase()
        self.buffer = SmartVideoBuffer(maxsize=3)
        
        # Session management
        self.session_counter = int(time.time())  # Unique session IDs
        self.current_session_id = None
        self.is_processing = False
        self.stop_event = threading.Event()
        self.video_thread = None
        
        # UI State
        self.photo_image = None
        self.after_id = None
        self.stats_history = deque(maxlen=100)  # For live charts
        
        self._setup_ui()
        self._start_update_loop()
        
    def _setup_ui(self):
        """Setup modern tabbed interface"""
        # Configure grid
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        # Create tabview
        if USE_CTK:
            self.tabview = ctk.CTkTabview(self.root)
        else:
            self.tabview = ttk.Notebook(self.root)
            
        self.tabview.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        # Create tabs
        if USE_CTK:
            self.tab_live = self.tabview.add("🔴 Live Detection")
            self.tab_history = self.tabview.add("📊 History & Analytics")
            self.tab_data = self.tabview.add("💾 Database View")
        else:
            self.tab_live = tk.Frame(self.tabview)
            self.tab_history = tk.Frame(self.tabview)
            self.tab_data = tk.Frame(self.tabview)
            self.tabview.add(self.tab_live, text="Live Detection")
            self.tabview.add(self.tab_history, text="History")
            self.tabview.add(self.tab_data, text="Database")
        
        self._setup_live_tab()
        self._setup_history_tab()
        self._setup_data_tab()
        
        # Status bar
        self.status_bar = ctk.CTkLabel(self.root, text="Ready", anchor="w") if USE_CTK else tk.Label(self.root, text="Ready", anchor="w", bd=1, relief=tk.SUNKEN)
        self.status_bar.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 5))
        
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)
        
    def _setup_live_tab(self):
        """Setup live detection interface"""
        # Main container
        main_frame = ctk.CTkFrame(self.tab_live) if USE_CTK else tk.Frame(self.tab_live)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Left: Video display
        video_frame = ctk.CTkFrame(main_frame) if USE_CTK else tk.LabelFrame(main_frame, text="Video Feed")
        video_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.video_canvas = tk.Canvas(video_frame, bg="black", highlightthickness=0)
        self.video_canvas.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.video_canvas.create_text(400, 300, text="Click 'Start Camera' or 'Upload Video'", 
                                     fill="white", font=("Helvetica", 16))
        
        # Right: Controls & Stats
        control_frame = ctk.CTkFrame(main_frame, width=350) if USE_CTK else tk.LabelFrame(main_frame, text="Controls", width=350)
        control_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=5, pady=5)
        control_frame.pack_propagate(False)
        
        # Control buttons
        if USE_CTK:
            btn_frame = ctk.CTkFrame(control_frame)
        else:
            btn_frame = tk.Frame(control_frame)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.btn_camera = ctk.CTkButton(btn_frame, text="📷 Start Camera", command=self._start_camera,
                                       fg_color="#2ecc71", height=40) if USE_CTK else tk.Button(btn_frame, text="Start Camera", command=self._start_camera, bg="#2ecc71", fg="white")
        self.btn_camera.pack(fill=tk.X, pady=2)
        
        self.btn_upload = ctk.CTkButton(btn_frame, text="📁 Upload Video", command=self._upload_video,
                                       fg_color="#3498db", height=40) if USE_CTK else tk.Button(btn_frame, text="Upload Video", command=self._upload_video, bg="#3498db", fg="white")
        self.btn_upload.pack(fill=tk.X, pady=2)
        
        self.btn_stop = ctk.CTkButton(btn_frame, text="⏹ Stop", command=self._stop_processing,
                                     fg_color="#e74c3c", height=40, state="disabled") if USE_CTK else tk.Button(btn_frame, text="Stop", command=self._stop_processing, bg="#e74c3c", fg="white", state="disabled")
        self.btn_stop.pack(fill=tk.X, pady=2)
        
        # Live statistics
        if USE_CTK:
            stats_frame = ctk.CTkFrame(control_frame)
        else:
            stats_frame = tk.LabelFrame(control_frame, text="Live Statistics")
        stats_frame.pack(fill=tk.X, padx=5, pady=10)
        
        self.stat_labels = {}
        stats = [
            ("Status", "Idle"),
            ("Session ID", "-"),
            ("Frame Count", "0"),
            ("Total Detections", "0"),
            ("Current FPS", "0"),
            ("Buffer Size", "0"),
            ("Excel Size", "0 MB")
        ]
        
        for name, default in stats:
            row = ctk.CTkFrame(stats_frame) if USE_CTK else tk.Frame(stats_frame)
            row.pack(fill=tk.X, pady=1)
            
            if USE_CTK:
                ctk.CTkLabel(row, text=f"{name}:", width=120, anchor="w").pack(side=tk.LEFT)
                lbl = ctk.CTkLabel(row, text=default, font=("Roboto", 12, "bold"))
            else:
                tk.Label(row, text=f"{name}:", width=15, anchor="w").pack(side=tk.LEFT)
                lbl = tk.Label(row, text=default, font=("Helvetica", 10, "bold"))
            lbl.pack(side=tk.LEFT)
            self.stat_labels[name] = lbl
        
        # Recent detections log
        if USE_CTK:
            log_frame = ctk.CTkFrame(control_frame)
        else:
            log_frame = tk.LabelFrame(control_frame, text="Recent Detections")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.detection_log = tk.Text(log_frame, height=10, width=30, font=("Consolas", 9))
        self.detection_log.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar = ttk.Scrollbar(log_frame, command=self.detection_log.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.detection_log.config(yscrollcommand=scrollbar.set)
        
        # Bottom charts
        chart_frame = ctk.CTkFrame(self.tab_live, height=200) if USE_CTK else tk.LabelFrame(self.tab_live, text="Live Analytics", height=200)
        chart_frame.pack(fill=tk.X, padx=10, pady=5, side=tk.BOTTOM)
        chart_frame.pack_propagate(False)
        
        self.fig, (self.ax1, self.ax2) = plt.subplots(1, 2, figsize=(10, 3))
        self.fig.tight_layout(pad=2.0)
        self.canvas = FigureCanvasTkAgg(self.fig, chart_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
    def _setup_history_tab(self):
        """Setup history and analytics view"""
        # Controls
        ctrl_frame = ctk.CTkFrame(self.tab_history) if USE_CTK else tk.Frame(self.tab_history)
        ctrl_frame.pack(fill=tk.X, padx=10, pady=5)
        
        if USE_CTK:
            ctk.CTkLabel(ctrl_frame, text="Date Range:").pack(side=tk.LEFT, padx=5)
            self.entry_start = ctk.CTkEntry(ctrl_frame, placeholder_text="YYYY-MM-DD", width=120)
            self.entry_start.pack(side=tk.LEFT, padx=5)
            self.entry_end = ctk.CTkEntry(ctrl_frame, placeholder_text="YYYY-MM-DD", width=120)
            self.entry_end.pack(side=tk.LEFT, padx=5)
            
            ctk.CTkButton(ctrl_frame, text="🔍 Analyze", command=self._analyze_range).pack(side=tk.LEFT, padx=5)
            ctk.CTkButton(ctrl_frame, text="📈 Create Charts", command=self._create_excel_charts).pack(side=tk.LEFT, padx=5)
            ctk.CTkButton(ctrl_frame, text="💾 Force Save", command=self._force_save).pack(side=tk.LEFT, padx=5)
        else:
            tk.Label(ctrl_frame, text="From:").pack(side=tk.LEFT, padx=5)
            self.entry_start = tk.Entry(ctrl_frame, width=12)
            self.entry_start.pack(side=tk.LEFT, padx=5)
            tk.Label(ctrl_frame, text="To:").pack(side=tk.LEFT, padx=5)
            self.entry_end = tk.Entry(ctrl_frame, width=12)
            self.entry_end.pack(side=tk.LEFT, padx=5)
            tk.Button(ctrl_frame, text="Analyze", command=self._analyze_range).pack(side=tk.LEFT, padx=5)
        
        # Set default dates
        today = datetime.now().strftime('%Y-%m-%d')
        self.entry_start.insert(0, today)
        self.entry_end.insert(0, today)
        
        # Analytics display
        display_frame = ctk.CTkFrame(self.tab_history) if USE_CTK else tk.Frame(self.tab_history)
        display_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Treeview for sessions
        columns = ("ID", "Date", "Start Time", "Source", "Frames", "Detections", "Classes", "Status")
        self.tree_sessions = ttk.Treeview(display_frame, columns=columns, show="headings", height=10)
        for col in columns:
            self.tree_sessions.heading(col, text=col)
            self.tree_sessions.column(col, width=100)
        
        scrollbar = ttk.Scrollbar(display_frame, orient="vertical", command=self.tree_sessions.yview)
        self.tree_sessions.configure(yscrollcommand=scrollbar.set)
        
        self.tree_sessions.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load initial data
        self._load_sessions()
        
    def _setup_data_tab(self):
        """Setup raw data view"""
        # Toolbar
        toolbar = ctk.CTkFrame(self.tab_data) if USE_CTK else tk.Frame(self.tab_data)
        toolbar.pack(fill=tk.X, padx=10, pady=5)
        
        if USE_CTK:
            ctk.CTkButton(toolbar, text="🔄 Refresh Data", command=self._refresh_data).pack(side=tk.LEFT, padx=5)
            ctk.CTkButton(toolbar, text="📋 Export CSV", command=self._export_csv).pack(side=tk.LEFT, padx=5)
            ctk.CTkButton(toolbar, text="🗑 Clear Old Data", command=self._clear_old_data).pack(side=tk.LEFT, padx=5)
        else:
            tk.Button(toolbar, text="Refresh", command=self._refresh_data).pack(side=tk.LEFT, padx=5)
            tk.Button(toolbar, text="Export CSV", command=self._export_csv).pack(side=tk.LEFT, padx=5)
        
        # Data display
        self.text_data = tk.Text(self.tab_data, wrap=tk.NONE, font=("Consolas", 9))
        self.text_data.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        scroll_y = ttk.Scrollbar(self.text_data, orient="vertical", command=self.text_data.yview)
        scroll_x = ttk.Scrollbar(self.text_data, orient="horizontal", command=self.text_data.xview)
        self.text_data.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        self._refresh_data()
        
    def _start_camera(self):
        """Start webcam"""
        self._start_processing(0, "camera")
        
    def _upload_video(self):
        """Upload video file"""
        filename = filedialog.askopenfilename(
            filetypes=[("Video files", "*.mp4 *.avi *.mov *.mkv"), ("All files", "*.*")]
        )
        if filename:
            self._start_processing(filename, "file")
            
    def _start_processing(self, source, source_type):
        """Initialize processing session"""
        if self.is_processing:
            return
        
        self.session_counter += 1
        self.current_session_id = self.session_counter
        self.is_processing = True
        self.stop_event.clear()
        
        # Update UI
        self.btn_camera.configure(state="disabled")
        self.btn_upload.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.stat_labels["Status"].configure(text="Running")
        self.stat_labels["Session ID"].configure(text=str(self.current_session_id))
        
        # Start database session
        self.db.start_session(self.current_session_id, source_type, str(source))
        
        # Clear buffer
        self.buffer = SmartVideoBuffer(maxsize=3)
        
        # Start video thread
        self.video_thread = VideoProcessor(
            self.model, self.buffer, self.stop_event, self.db,
            self.current_session_id, source_type, source
        )
        self.video_thread.start()
        
        self.status_bar.configure(text=f"Session {self.current_session_id} started - Recording to Excel...")
        
    def _stop_processing(self):
        """Stop current session"""
        if not self.is_processing:
            return
        
        self.stop_event.set()
        
        if self.video_thread and self.video_thread.is_alive():
            self.video_thread.join(timeout=3.0)
        
        self.is_processing = False
        self.current_session_id = None
        
        # Update UI
        self.btn_camera.configure(state="normal")
        self.btn_upload.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        self.stat_labels["Status"].configure(text="Stopped")
        
        self.video_canvas.delete("all")
        self.video_canvas.create_text(400, 300, text="Stopped - Data saved to Excel", 
                                     fill="white", font=("Helvetica", 16))
        
        self.status_bar.configure(text="Session ended - Data saved")
        
        # Refresh views
        self._load_sessions()
        self._refresh_data()
        
    def _start_update_loop(self):
        """Start UI update loop"""
        self._update_frame()
        self._update_stats()
        
    def _update_frame(self):
        """Update video frame"""
        if self.is_processing:
            data = self.buffer.get(timeout=0.001)
            if data:
                frame, detections, fps, timestamp = data
                self._display_frame(frame)
                self._update_detection_log(detections, timestamp)
                self.stats_history.append({
                    'time': timestamp,
                    'detections': len(detections),
                    'fps': fps
                })
        
        self.after_id = self.root.after(33, self._update_frame)  # ~30 FPS display
    
    def _display_frame(self, frame):
        """Display frame on canvas"""
        try:
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            
            # Resize to fit canvas
            canvas_w = self.video_canvas.winfo_width()
            canvas_h = self.video_canvas.winfo_height()
            
            if canvas_w > 1 and canvas_h > 1:
                h, w = frame_rgb.shape[:2]
                scale = min(canvas_w/w, canvas_h/h)
                new_w, new_h = int(w*scale), int(h*scale)
                frame_rgb = cv2.resize(frame_rgb, (new_w, new_h))
            
            img = Image.fromarray(frame_rgb)
            self.photo_image = ImageTk.PhotoImage(image=img)
            
            self.video_canvas.delete("all")
            self.video_canvas.create_image(canvas_w//2, canvas_h//2, image=self.photo_image, anchor=tk.CENTER)
            
        except Exception as e:
            logger.error(f"Display error: {e}")
    
    def _update_detection_log(self, detections, timestamp):
        """Update detection log"""
        if detections:
            time_str = timestamp.strftime('%H:%M:%S')
            for det in detections:
                log_entry = f"[{time_str}] {det['class']}: {det['confidence']:.2f}\n"
                self.detection_log.insert(tk.END, log_entry)
                self.detection_log.see(tk.END)
    
    def _update_stats(self):
        """Update statistics display"""
        if self.is_processing and self.video_thread:
            self.stat_labels["Frame Count"].configure(text=str(self.video_thread.frame_count))
            self.stat_labels["Total Detections"].configure(text=str(self.video_thread.total_detections))
            
            # Get buffer stats
            buf_stats = self.db.get_live_stats()
            self.stat_labels["Buffer Size"].configure(text=str(buf_stats['buffered_records']))
            self.stat_labels["Excel Size"].configure(text=f"{buf_stats['file_size_mb']} MB")
            
            if self.stats_history:
                latest = self.stats_history[-1]
                self.stat_labels["Current FPS"].configure(text=f"{latest['fps']:.1f}")
        
        # Update charts every second
        if hasattr(self, '_chart_counter'):
            self._chart_counter += 1
        else:
            self._chart_counter = 0
        
        if self._chart_counter % 30 == 0:  # Every ~1 second
            self._update_live_charts()
        
        self.root.after(1000, self._update_stats)  # Update every second
    
    def _update_live_charts(self):
        """Update live analytics charts"""
        try:
            self.ax1.clear()
            self.ax2.clear()
            
            if len(self.stats_history) > 1:
                times = [i for i in range(len(self.stats_history))]
                detections = [s['detections'] for s in self.stats_history]
                fps_values = [s['fps'] for s in self.stats_history]
                
                # Detection timeline
                self.ax1.plot(times, detections, marker='o', color='#3498db', linewidth=2)
                self.ax1.fill_between(times, detections, alpha=0.3, color='#3498db')
                self.ax1.set_title("Detections per Frame")
                self.ax1.set_xlabel("Frame")
                
                # FPS timeline
                self.ax2.plot(times, fps_values, marker='s', color='#2ecc71', linewidth=2)
                self.ax2.axhline(y=30, color='r', linestyle='--', alpha=0.5, label='Target 30 FPS')
                self.ax2.set_title("Processing FPS")
                self.ax2.set_xlabel("Frame")
                self.ax2.legend()
            
            self.fig.tight_layout()
            self.canvas.draw()
            
        except Exception as e:
            logger.error(f"Chart update error: {e}")
    
    def _load_sessions(self):
        """Load session history from Excel"""
        for item in self.tree_sessions.get_children():
            self.tree_sessions.delete(item)
        
        try:
            df = pd.read_excel(self.db.filename, sheet_name="Sessions")
            if not df.empty:
                for _, row in df.iterrows():
                    self.tree_sessions.insert("", tk.END, values=(
                        row['Session_ID'],
                        row['Date'],
                        row['Start_Time'],
                        f"{row['Source_Type']}: {str(row['Source_Path'])[:20]}...",
                        row['Total_Frames'],
                        row['Total_Detections'],
                        row['Unique_Classes'],
                        row['Status']
                    ))
        except Exception as e:
            logger.error(f"Error loading sessions: {e}")
    
    def _analyze_range(self):
        """Analyze date range"""
        start = self.entry_start.get()
        end = self.entry_end.get()
        
        try:
            df = self.db.get_date_range_data(start, end)
            if df.empty:
                messagebox.showinfo("Info", "No data found for selected range")
                return
            
            # Show summary
            summary = f"""
Date Range Analysis: {start} to {end}
Total Records: {len(df)}
Unique Classes: {df['Class_Name'].nunique()}
Most Common: {df['Class_Name'].mode().iloc[0] if not df['Class_Name'].mode().empty else 'N/A'}
Average Confidence: {df['Confidence'].mean():.2%}
            """
            messagebox.showinfo("Analysis Results", summary)
            
        except Exception as e:
            messagebox.showerror("Error", f"Analysis failed: {e}")
    
    def _create_excel_charts(self):
        """Create charts in Excel file"""
        self.db.create_analytics_charts()
        messagebox.showinfo("Success", "Charts created in Excel file!")
    
    def _force_save(self):
        """Force immediate save to Excel"""
        self.db.force_flush()
        messagebox.showinfo("Success", "Data flushed to Excel!")
    
    def _refresh_data(self):
        """Refresh raw data view"""
        self.text_data.delete(1.0, tk.END)
        try:
            df = self.db.read_all_data()
            self.text_data.insert(tk.END, f"Total Records: {len(df)}\n")
            self.text_data.insert(tk.END, f"File: {self.db.filename}\n")
            self.text_data.insert(tk.END, "="*80 + "\n")
            self.text_data.insert(tk.END, df.tail(50).to_string())  # Show last 50 records
        except Exception as e:
            self.text_data.insert(tk.END, f"Error reading data: {e}")
    
    def _export_csv(self):
        """Export to CSV"""
        filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if filename:
            try:
                df = self.db.read_all_data()
                df.to_csv(filename, index=False)
                messagebox.showinfo("Success", f"Exported to {filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Export failed: {e}")
    
    def _clear_old_data(self):
        """Clear data older than specified days"""
        days = simpledialog.askinteger("Clear Data", "Delete data older than how many days?", initialvalue=30)
        if days:
            # Implementation would go here
            messagebox.showinfo("Info", "Feature: Clear old data (implementation pending)")
    
    def _on_closing(self):
        """Clean shutdown"""
        if self.is_processing:
            self._stop_processing()
        
        # Final flush
        self.db.force_flush()
        
        if self.after_id:
            self.root.after_cancel(self.after_id)
        
        self.root.destroy()
        logger.info("Application closed")
    
    def run(self):
        self.root.mainloop()


def check_dependencies():
    """Check and report dependencies"""
    deps = {
        'cv2': 'opencv-python',
        'numpy': 'numpy',
        'PIL': 'pillow',
        'matplotlib': 'matplotlib',
        'pandas': 'pandas',
        'openpyxl': 'openpyxl'
    }
    
    missing = []
    for module, package in deps.items():
        try:
            __import__(module)
            print(f"✓ {package}")
        except ImportError:
            missing.append(package)
            print(f"✗ {package} (missing)")
    
    if missing:
        print(f"\nInstall missing packages: pip install {' '.join(missing)}")
        return False
    return True


def main():
    print("="*60)
    print("Advanced Object Detection System - Excel Live Database")
    print("Features: Real-time Excel writing, Persistent storage, Modern UI")
    print("="*60)
    
    if not check_dependencies():
        return
    
    app = ModernDetectionGUI()
    app.run()


if __name__ == "__main__":
    main()